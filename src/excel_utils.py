import re
import pandas as pd
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell

# ==========================================
# DEFENSIVE ABSTRACTION LAYER
# ==========================================
def safe_write(sheet, r, c, val):
    """
    Safely writes a value to an openpyxl cell.
    If the cell is a secondary component of a MergedCell, it safely ignores it
    to prevent 'read-only' XML crashes.
    """
    cell = sheet.cell(row=r, column=c)
    if not isinstance(cell, MergedCell):
        cell.value = val

# ==========================================
# 1. DYNAMIC TABLE UTILITIES (E.g., Jadual 1, 3, 5, 6)
# ==========================================
def get_dynamic_boundaries(sheet, header_text):
    """Dynamically finds the table boundaries using openpyxl syntax."""
    target_row = None
    for row_idx in range(1, 100):  
        val = sheet.cell(row=row_idx, column=2).value 
        if val and isinstance(val, str) and header_text.lower() in val.lower():
            target_row = row_idx
            break
            
    if not target_row: return None, None, None, None
        
    start_col = None
    end_col = None
    
    for col_idx in range(3, 30):  
        val = sheet.cell(row=target_row, column=col_idx).value
        if val and isinstance(val, str):
            clean_val = val.strip()
            if re.match(r"^P\.\d+", clean_val, re.IGNORECASE) and start_col is None:
                start_col = col_idx
            if re.match(r"^N\.\d+", clean_val, re.IGNORECASE):
                end_col = col_idx 
                
    if start_col and not end_col: end_col = start_col
    if start_col and end_col:
        max_slots = (end_col - start_col) + 1
        return target_row, start_col, end_col, max_slots
        
    return target_row, None, None, None

def inject_dynamic_table(sheet, target_row, start_col, max_slots, locations_to_inject, row_map, max_row_clean, hierarchy):
    """Handles right-aligned math, safe template cleanup, and dynamic injection."""
    payload = min(len(locations_to_inject), max_slots)
    locations_to_inject = locations_to_inject[:max_slots]
    skip_offset = max_slots - payload
    actual_start_col = start_col + skip_offset

    # SAFELY Clean left-side dummy data (This is where the crash originally happened!)
    if skip_offset > 0:
        for r in range(target_row, max_row_clean + 1):
            for c in range(start_col, actual_start_col):
                safe_write(sheet, r, c, None)

    for i, (loc_code, loc_name, metrics_data) in enumerate(locations_to_inject):
        current_col = actual_start_col + i
        
        # Inject Header
        safe_write(sheet, target_row, current_col, f"{loc_code}\n{loc_name}")
        
        if loc_code in [hierarchy.get('parl_code'), hierarchy.get('parent_parl_code')]:
            cell = sheet.cell(row=target_row, column=current_col)
            if not isinstance(cell, MergedCell):
                cell.font = Font(bold=True)
        
        # Inject Data Rows
        for row_idx, metric_info in row_map.items():
            if isinstance(metric_info, tuple):
                metric_name, year = metric_info
                year_data = metrics_data.get(str(year), {})
            else:
                metric_name = metric_info
                year_data = metrics_data 
                
            val = year_data.get(metric_name, "n.a")
            
            if pd.notna(val) and val not in ["n.a", ""]:
                try: val = float(val)
                except ValueError: pass
            else:
                val = "n.a"
                
            safe_write(sheet, row_idx, current_col, val)

# ==========================================
# 2. STATIC TABLE UTILITIES (E.g., Jadual 2.0, 2.2, 4.0)
# ==========================================
def inject_static_table(sheet, metrics_data, row_map, col_map):
    """
    Injects data into fixed-grid tables where rows are metrics and columns are years.
    """
    for col_idx, year in col_map.items():
        year_data = metrics_data.get(str(year), {})
        
        for row_idx, metric_info in row_map.items():
            # Support both flat strings and tuples just in case
            metric_name = metric_info[0] if isinstance(metric_info, tuple) else metric_info
                
            val = year_data.get(metric_name, "n.a")
            
            if pd.notna(val) and val not in ["n.a", ""]:
                try: val = float(val)
                except ValueError: pass
            else:
                val = "n.a"
                
            safe_write(sheet, row_idx, col_idx, val)